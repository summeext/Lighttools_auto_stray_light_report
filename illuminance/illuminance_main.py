"""
auto_stray_light — Stray Light Auto Analysis (照度版 / Illuminance)
Python COM API + tkinter GUI

与强度版的区别:
- 数据源: ILLUMINANCE_MESH (照度) 替代 INTENSITY_MESH (强度)
- 图表: Forward Illuminance 替代 Forward Intensity
- 图表输出为 BMP 格式 (避免 PNG 被加密软件自动加密导致无法读取)
"""

import os, csv, time, sys, threading
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import ttk, filedialog, messagebox


def t(result):
    if isinstance(result, tuple) and len(result) == 2:
        return result[0], result[1]
    return result, None


def clean_output(output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)
    for pattern in ["*.csv", "*.CSV", "*.png", "*.PNG", "*.bmp", "*.BMP",
                     "*.jpg", "*.JPG", "*.xlsx", "*.XLSX", "*.txt"]:
        for f in output_dir.glob(pattern):
            f.unlink()
    print("[Step 0] Output cleaned")


def wait_stable(lt, imesh_key, max_wait=15.0, interval=0.1):
    """Poll MAX VALUE until stable (two consecutive reads equal) or timeout."""
    prev = None
    for _ in range(int(max_wait / interval)):
        time.sleep(interval)
        cur, _ = t(lt.DbGet(imesh_key, "MAX VALUE"))
        if cur is not None and prev is not None and cur == prev:
            return cur
        prev = cur
    return prev  # timeout, return last read


def write_macros(project_dir, output_dir):
    """Generate .ltb files with absolute paths filled from project_dir."""
    proj_d = str(project_dir).replace("\\", "\\\\")
    out_d = str(output_dir).replace("\\", "\\\\")

    (project_dir / "illumChart.1.ltb").write_text(
        'REM illumChart.1.ltb — auto-generated\n'
        'LTBSET UPDATE 0\n'
        'LTBSET DIALOG 0\n'
        f'OPEN "{proj_d}\\\\save_chart_params.txt" FOR INPUT AS #1\n'
        'INPUT #1, filePath$\n'
        'CLOSE #1\n'
        'LTCMD "PrintToFile" + " " + LTSTR$(filePath$)\n'
        'LTBSET UPDATE 1\n'
        'END\n'
    )


def main(project_dir=None, receiver_name="Receiver_38", b_percent=0.005, max_paths=50):
    import win32com.client

    if project_dir is None:
        project_dir = Path(__file__).parent
    else:
        project_dir = Path(project_dir)
    output_dir = project_dir / "output"

    fso_path = (
        'ILLUM_MANAGER[Illumination Manager].RECEIVERS[Receiver List]'
        f'.SURFACE_RECEIVER[{receiver_name}]'
        '.FORWARD_SIM_FUNCTION[Forward Simulation]'
    )

    # Ensure LightTools can find helper macros in project directory
    macro_dir = str(project_dir)
    existing = os.environ.get("ORA_SCRIPT_PATH", "")
    if macro_dir not in existing:
        sep = ";" if existing else ""
        os.environ["ORA_SCRIPT_PATH"] = macro_dir + sep + existing

    lt = win32com.client.Dispatch("LightTools.LTAPI")
    lt.Begin()

    write_macros(project_dir, output_dir)
    clean_output(output_dir)

    # ---- Step 1: Setup flags ----
    sim_list, _ = t(lt.DbList("ILLUM_MANAGER[1]", "SIMULATIONS"))
    sim_key, _ = t(lt.ListAtPos(sim_list, 1))
    lt.DbSet(sim_key, "COLLECTALLFORWARDRAYPATHS", "Yes")
    lt.DbSet(sim_key, "COLLECTRAYPATHS", "Yes")

    rec_list, _ = t(lt.DbList("ILLUM_MANAGER[1]", "RECEIVER"))
    rec_key, _ = t(lt.ListAtPos(rec_list, 1))
    fsf_list, _ = t(lt.DbList(rec_key, "FORWARD_SIM_FUNCTION"))
    fsf_key, _ = t(lt.ListAtPos(fsf_list, 1))
    lt.DbSet(fsf_key, "COLLECTRAYPATHS", "Yes")
    lt.DbSet(fsf_key, "SHOWRAYPATHS", "Yes")
    print("[Step 1] Flags set")

    # ---- Step 2: Compute filtered mesh data for Ray Path analysis ----
    lt.Cmd("RayPath")
    lt.Cmd("RayPath")
    lt.Cmd(f'\\O"{fso_path}"')
    lt.Cmd("Compute=")
    lt.Cmd("\\Q")
    time.sleep(1)
    print("[Step 2] RayPath Compute done")

    # ---- Step 2a: Read path detail strings via COM (bypasses macro 999-char limit) ----
    path_strings = []
    try:
        npaths_val, _ = t(lt.DbGet(fsf_key, "NumberOfRayPaths"))
        npaths_pre = int(npaths_val) if npaths_val else 0
        status = 0
        for i in range(1, npaths_pre + 1):
            s, _ = t(lt.DbGet(fsf_key, "RayPathStringAt", status, i))
            # COM returns lines separated by \r\n; normalize to \n
            path_strings.append(s.replace("\r\n", "\n") if s else "")
        non_empty = sum(1 for s in path_strings if s.strip())
        max_len = max((len(s) for s in path_strings), default=0)
        print(f"[Step 2a] Path details via COM: {non_empty} non-empty, max {max_len} chars")
    except Exception as e:
        print(f"[Step 2a] COM export failed: {e}")

    # ---- Re-find FSF ----
    lt.ListDelete(fsf_list)
    lt.ListDelete(rec_list)
    rec_list, _ = t(lt.DbList("ILLUM_MANAGER[1]", "RECEIVER"))
    rec_key, _ = t(lt.ListAtPos(rec_list, 1))
    fsf_list, _ = t(lt.DbList(rec_key, "FORWARD_SIM_FUNCTION"))
    fsf_key, _ = t(lt.ListAtPos(fsf_list, 1))
    imesh_list, _ = t(lt.DbList(fsf_key, "ILLUMINANCE_MESH"))
    imesh_key, _ = t(lt.ListAtPos(imesh_list, 1))
    npaths_val, _ = t(lt.DbGet(fsf_key, "NumberOfRayPaths"))
    npaths = int(npaths_val) if npaths_val else 0
    print(f"[Step 3] {npaths} ray paths found")

    # ---- Step 3b: Read Max Illuminance per path ----
    path_data = []
    for i in range(1, npaths + 1):
        # HideAll + show only path i
        lt.Cmd(f'\\O"{fso_path}"')
        lt.Cmd("HideAll=")
        lt.Cmd("\\Q")
        time.sleep(0.15)

        lt.Cmd(f'\\O"{fso_path}"')
        lt.Cmd(f"RayPathVisibleAt[{i}]=Yes")
        lt.Cmd("\\Q")

        max_val = wait_stable(lt, imesh_key)
        # Get path detail from macro-exported strings (0-based array)
        pstr = path_strings[i - 1] if i - 1 < len(path_strings) else ""
        path_data.append({
            "path_number": i,
            "max_illuminance": float(max_val) if max_val else 0.0,
            "path_detail": pstr,
        })
        print(f"  Path {i}: {path_data[-1]['max_illuminance']:.6f}")

    # ---- Step 4: Sort descending ----
    path_data.sort(key=lambda p: -p["max_illuminance"])
    top = path_data[0]["max_illuminance"]
    threshold = top * (b_percent / 100.0)
    print(f"[Step 4] Max: {top:.8f}, Threshold: {threshold:.8f}")

    # ---- Step 5: Export review.csv ----
    filtered = [p for p in path_data if p["max_illuminance"] >= threshold]
    if len(filtered) > max_paths:
        print(f"[Step 4] Truncating from {len(filtered)} to {max_paths} paths")
        filtered = filtered[:max_paths]
    with open(output_dir / "review.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Path Number", "Max Illuminance"])
        for p in filtered:
            w.writerow([p["path_number"], f"{p['max_illuminance']:.8f}"])
    print(f"[Step 5] review.csv: {len(filtered)} paths")

    # ---- Step 6: Individual path detail CSVs ----
    for p in filtered:
        fname = output_dir / f"Path_{p['path_number']}.csv"
        with open(fname, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["Path Number", "Max Illuminance", "Path Detail"])
            w.writerow([p["path_number"], f"{p['max_illuminance']:.8f}", p["path_detail"]])
    print(f"[Step 6] {len(filtered)} path detail CSVs saved")

    # ---- Step 7: Save Forward Illuminance chart images (BMP format) ----
    for p in filtered:
        # Isolate path (Python)
        lt.Cmd(f'\\O"{fso_path}"')
        lt.Cmd("HideAll=")
        lt.Cmd("\\Q")
        time.sleep(0.2)

        lt.Cmd(f'\\O"{fso_path}"')
        lt.Cmd(f"RayPathVisibleAt[{p['path_number']}]=Yes")
        lt.Cmd("\\Q")

        # Re-acquire mesh key (invalidated by previous chart switch)
        s7_rec_list, _ = t(lt.DbList("ILLUM_MANAGER[1]", "RECEIVER"))
        s7_rec_key, _ = t(lt.ListAtPos(s7_rec_list, 1))
        s7_fsf_list, _ = t(lt.DbList(s7_rec_key, "FORWARD_SIM_FUNCTION"))
        s7_fsf_key, _ = t(lt.ListAtPos(s7_fsf_list, 1))
        s7_imesh_list, _ = t(lt.DbList(s7_fsf_key, "ILLUMINANCE_MESH"))
        s7_imesh_key, _ = t(lt.ListAtPos(s7_imesh_list, 1))
        wait_stable(lt, s7_imesh_key)
        lt.ListDelete(s7_imesh_list)
        lt.ListDelete(s7_fsf_list)
        lt.ListDelete(s7_rec_list)

        # Switch to chart view (Python)
        lt.Cmd(f"\\VChart_{receiver_name}_Forward_Illuminance")
        time.sleep(0.5)

        # Save image via minimal macro — use BMP to avoid encryption issues
        img_file = str(output_dir / f"Path_{p['path_number']}.bmp")
        params_file = project_dir / "save_chart_params.txt"
        params_file.write_text(f"{img_file}\n", encoding="utf-8")
        lt.Cmd("illumChart.1.ltb")
        time.sleep(0.5)

        # Back to 3D
        lt.Cmd("\\V3D")
        print(f"  Image saved: Path_{p['path_number']}.bmp")
    print(f"[Step 7] {len(filtered)} chart images saved")

    # ---- Step 8: Generate Excel report ----
    from PIL import Image as PILImage

    wb = Workbook()
    header_font = Font(name="Arial", bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    diff_font = Font(name="Arial", color="FF0000", bold=True, size=10)  # red bold
    normal_font = Font(name="Arial", color="000000", size=10)
    ref_font = Font(name="Arial", color="333333", size=10)         # grey for reference
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    # Main ray path = first in sorted order (highest illuminance)
    main_path = filtered[0]
    main_lines = main_path["path_detail"].strip().split("\n")
    main_lines = [l.strip() for l in main_lines if l.strip()]

    # --- Review sheet ---
    ws_review = wb.active
    ws_review.title = "Review"
    main_illuminance = main_path["max_illuminance"]
    ws_review.append(["Path Number", "Max Illuminance", "Illuminance Ratio", "Note"])
    for cell in ws_review[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    for p in filtered:
        ratio = p["max_illuminance"] / main_illuminance if main_illuminance else 0
        note = "Main Ray Path" if p is main_path else ""
        ws_review.append([p["path_number"], p["max_illuminance"], ratio, note])
        for cell in ws_review[ws_review.max_row]:
            cell.border = thin_border
        ws_review.cell(row=ws_review.max_row, column=2).number_format = "0.00000000"
        ws_review.cell(row=ws_review.max_row, column=3).number_format = "0.00%"
    ws_review.column_dimensions["A"].width = 14
    ws_review.column_dimensions["B"].width = 18
    ws_review.column_dimensions["C"].width = 12
    ws_review.column_dimensions["D"].width = 20

    # --- Path detail sheets ---
    for p in filtered:
        sheet_name = f"Path_{p['path_number']}"
        ws = wb.create_sheet(sheet_name)
        is_main = (p is main_path)

        # Row 1: Path info
        ratio = p["max_illuminance"] / main_illuminance if main_illuminance else 0
        ws.merge_cells("A1:D1")
        c = ws["A1"]
        c.value = f"Path {p['path_number']}  |  Max Illuminance: {p['max_illuminance']:.8f}  |  Illuminance Ratio: {ratio:.2%}"
        if is_main:
            c.value += "  (Main Ray Path — Reference)"
        c.font = Font(name="Arial", bold=True, size=13, color="1F4E79")
        c.alignment = Alignment(horizontal="left")

        # Row 3: Column headers for detail comparison
        ws.append([])  # row 2 blank
        ws.append(["#", f"Path {p['path_number']} Surface Sequence", "Main Path (Path {main_path['path_number']}) Reference"])
        for cell in ws[3]:
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.fill = header_fill
            cell.border = thin_border

        # Split detail into lines
        detail_lines = p["path_detail"].strip().split("\n")
        detail_lines = [l.strip() for l in detail_lines if l.strip()]

        max_lines = max(len(detail_lines), len(main_lines))
        diff_count = 0
        for i in range(max_lines):
            row = i + 4
            cur = detail_lines[i] if i < len(detail_lines) else ""
            ref = main_lines[i] if i < len(main_lines) else ""
            is_different = (cur != ref)

            ws.cell(row=row, column=1, value=i + 1).border = thin_border
            ws.cell(row=row, column=1).font = normal_font

            cell_cur = ws.cell(row=row, column=2, value=cur)
            cell_cur.border = thin_border
            cell_cur.font = diff_font if is_different else normal_font

            cell_ref = ws.cell(row=row, column=3, value=ref)
            cell_ref.border = thin_border
            cell_ref.font = ref_font
            if is_different:
                diff_count += 1

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 55

        # Embed image (BMP format) with preserved aspect ratio
        img_row = max_lines + 6
        bmp_file = output_dir / f"Path_{p['path_number']}.bmp"
        if bmp_file.exists():
            with PILImage.open(str(bmp_file)) as pil_img:
                orig_w, orig_h = pil_img.size
            max_w = 500
            scale = max_w / orig_w
            new_w = orig_w * scale
            new_h = orig_h * scale

            img = XLImage(str(bmp_file))
            img.width = new_w
            img.height = new_h
            ws.add_image(img, f"A{img_row}")

        if not is_main:
            print(f"  Path_{p['path_number']}: {diff_count} lines differ from main path")

    # Save
    xlsx_path = output_dir / "stray_light_report_illuminance.xlsx"
    wb.save(str(xlsx_path))
    print(f"[Step 8] Excel report saved: {xlsx_path.name}")

    # ---- Cleanup ----
    for f in [project_dir / "save_chart_params.txt"]:
        if f.exists():
            f.unlink()
    lt.ListDelete(imesh_list)
    lt.ListDelete(fsf_list)
    lt.ListDelete(rec_list)
    lt.ListDelete(sim_list)
    lt.End()
    print("[Completed]\n\n· Author: Xu Tian ·")


class RedirectText:
    """Redirect stdout to a tkinter Text widget (thread-safe via queue)."""
    def __init__(self, text_widget):
        self.text_widget = text_widget
        self.queue = []
    def write(self, s):
        self.queue.append(s)
    def flush(self):
        while self.queue:
            self.text_widget.insert(tk.END, self.queue.pop(0))
            self.text_widget.see(tk.END)
            self.text_widget.update_idletasks()


class App:
    def __init__(self, root):
        self.root = root
        root.title("Stray Light Auto Analysis (Illuminance)")
        root.resizable(True, True)
        root.geometry("680x520")

        frame = ttk.Frame(root, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        # Project dir
        ttk.Label(frame, text="Project Dir:").grid(row=0, column=0, sticky="w", pady=2)
        self.dir_var = tk.StringVar(value=str(Path(__file__).parent))
        ttk.Entry(frame, textvariable=self.dir_var, width=55).grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(frame, text="Browse...", command=self._browse).grid(row=0, column=2)

        # Receiver name
        ttk.Label(frame, text="Receiver Name:").grid(row=1, column=0, sticky="w", pady=2)
        self.recv_var = tk.StringVar(value="Receiver_38")
        ttk.Entry(frame, textvariable=self.recv_var, width=55).grid(row=1, column=1, columnspan=2, sticky="ew", padx=5)

        # B% threshold
        ttk.Label(frame, text="B% Threshold:").grid(row=2, column=0, sticky="w", pady=2)
        self.b_var = tk.StringVar(value="0.005")
        ttk.Entry(frame, textvariable=self.b_var, width=55).grid(row=2, column=1, columnspan=2, sticky="ew", padx=5)

        # Max paths
        ttk.Label(frame, text="Max Paths:").grid(row=3, column=0, sticky="w", pady=2)
        self.maxp_var = tk.StringVar(value="50")
        ttk.Entry(frame, textvariable=self.maxp_var, width=55).grid(row=3, column=1, columnspan=2, sticky="ew", padx=5)

        # Run button
        self.run_btn = ttk.Button(frame, text="Run", command=self._run)
        self.run_btn.grid(row=4, column=0, columnspan=3, pady=10)

        # Log output
        self.log = tk.Text(frame, wrap=tk.WORD, bg="#1E1E1E", fg="#D4D4D4",
                           insertbackground="white", font=("Consolas", 9))
        self.log.grid(row=5, column=0, columnspan=3, sticky="nsew", pady=(5, 0))
        frame.rowconfigure(5, weight=1)
        self.redirect = RedirectText(self.log)
        frame.columnconfigure(1, weight=1)

    def _browse(self):
        d = filedialog.askdirectory(title="Select project directory")
        if d:
            self.dir_var.set(d)

    def _run(self):
        self.run_btn.config(state=tk.DISABLED, text="Running...")
        self.log.delete("1.0", tk.END)
        sys.stdout = self.redirect

        def worker():
            try:
                main(
                    project_dir=self.dir_var.get(),
                    receiver_name=self.recv_var.get(),
                    b_percent=float(self.b_var.get()),
                    max_paths=int(self.maxp_var.get()),
                )
            except Exception as e:
                print(f"\nERROR: {e}")

            self.root.after(0, self._done)

        threading.Thread(target=worker, daemon=True).start()
        self.root.after(100, self._poll_log)

    def _poll_log(self):
        self.redirect.flush()
        if self.run_btn["state"] == tk.DISABLED:
            self.root.after(100, self._poll_log)
        else:
            self.redirect.flush()

    def _done(self):
        sys.stdout = sys.__stdout__
        self.redirect.flush()
        self.run_btn.config(state=tk.NORMAL, text="Run")
        messagebox.showinfo("Done", "Stray light analysis completed.")


if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
